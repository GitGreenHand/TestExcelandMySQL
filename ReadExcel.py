# coding= utf-8
from openpyxl import load_workbook
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
# 加载要读取的文件
try:
	wb = load_workbook("test.xlsx")
except:
	print('没找到该文件')  # 获取当前的操作页面

sheet = wb.get_active_sheet()
# 获取当前及界面的行最大值和列最大值
# 行
max_col = sheet.max_column
# 列
max_row = sheet.max_row
# 通过双循环遍历其中的值
# 将获取的数据写入到文件中
f= open ("excel.txt",'w')
for row in range(1, max_row+1):
	for col in range(1, max_col+1):
		content=sheet.cell(row, col)
		f.write(str(content.value)+',')
	f.write('\n')
f.close()



"""

# 默认可读写，若有需要可以指定write_only和read_only为True
wb = load_workbook('text.xlsx')

# 获得所有sheet的名称
print(wb.get_sheet_names())
# 根据sheet名字获得sheet
a_sheet = wb.get_sheet_by_name('Sheet1')
# 获得sheet名
print(a_sheet.title)
# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
sheet = wb.active

# 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
b4 = sheet['B4']
# 分别返回
print(f'({b4.column}, {b4.row}) is {b4.value}')  # 返回的数字就是int型

# 除了用下标的方式获得，还可以用cell函数, 换成数字，这个表示B4
b4_too = sheet.cell(row=4, column=2)
print(b4_too.value)

# 获得最大列和最大行
print(sheet.max_row)
print(sheet.max_column)

# 因为按行，所以返回A1, B1, C1这样的顺序
for row in sheet.rows:
    for cell in row:
        print(cell.value)

# A1, A2, A3这样的顺序
for column in sheet.columns:
    for cell in column:
        print(cell.value)
        """
